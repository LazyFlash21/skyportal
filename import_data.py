import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'skyportal.settings')
django.setup()

from openpyxl import load_workbook
from core.models import Department, Team

# Clear existing data
Team.objects.all().delete()
Department.objects.all().delete()

wb = load_workbook('data/Agile Project Module UofW - Team Registry.xlsx')
ws = wb.active

rows = list(ws.iter_rows(values_only=True))

for row in rows[1:]:  # skip header
    if not row[0]:
        continue

    dept_name = row[0]
    dept_head = row[2] if row[2] else ''

    # Get or create department
    dept, created = Department.objects.get_or_create(
        name=dept_name,
        defaults={'head': dept_head}
    )

    # Create team
    Team.objects.create(
        department=dept,
        name=row[3] if row[3] else '',
        leader=row[1] if row[1] else '',
        jira_project=row[4] if row[4] else '',
        github_repo=row[6] if row[6] else '',
        jira_board=row[7] if row[7] else '',
        development_focus=row[8] if row[8] else '',
        key_skills=row[9] if row[9] else '',
        downstream_dependencies=row[10] if row[10] else '',
        dependency_type=row[11] if row[11] else '',
        slack_channels=row[15] if row[15] else '',
        standup_time=row[16] if row[16] else '',
        agile_practices=row[17] if row[17] else '',
        team_wiki=row[18] if row[18] else '',
        concurrent_projects=str(row[19]) if row[19] else '',
    )

print(f'Done! Imported {Team.objects.count()} teams across {Department.objects.count()} departments')